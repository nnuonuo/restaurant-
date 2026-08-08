import streamlit as st
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

st.set_page_config(page_title="Restaurant Inventory & Chain Management System", layout="wide")

# --- CSS TÙY CHỈNH GIAO DIỆN ---
st.markdown("""
    <style>
    .stApp {
        background-color: #f7f9fa;
    }
    [data-testid="stSidebar"] {
        background-color: #b4cfdc;
        border-right: 1px solid #90bcd5;
    }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3, [data-testid="stSidebar"] span, [data-testid="stSidebar"] label {
        color: #1e293b !important;
    }
    div.stButton > button:first-child, div.stFormSubmitButton > button:first-child {
        background-color: #bed650 !important;
        color: #1e293b !important;
        font-weight: bold;
        border: 1px solid #a8c238;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        transition: all 0.3s ease;
    }
    div.stButton > button:first-child:hover, div.stFormSubmitButton > button:first-child:hover {
        background-color: #a8c238 !important;
        border-color: #96af29;
    }
    .stTextInput input, .stSelectbox select, .stNumberInput input {
        border-color: #90bcd5 !important;
        border-radius: 6px;
    }
    h1 {
        color: #2c3e50;
    }
    .stAlert {
        border-radius: 8px;
    }
    </style>
""", unsafe_allow_html=True)

DATA_FILE = "restaurant_inventory_data.xlsx"
EXPORT_FILE = "branch_export_history.csv"
TRANSFER_FILE = "branch_transfer_history.csv"
ORDER_FILE = "branch_order_requests.csv"
SECRET_ACTION_PWD = "264221"

# --- QUẢN LÝ SESSION STATE ---
if "passwords" not in st.session_state:
    st.session_state.passwords = {
        "nuonuo": "264221",
        "heni": "Heni2026",
        "admin": "budapest2026",
        "shibuya": "shibuya123",
        "geisha.baross": "geisha2023",
        "geisha.corvin": "Corvin2026",
        "urbn": "ub2026",
        "matchy": "matchy2026"
    }

if "reset_step" not in st.session_state:
    st.session_state.reset_step = 1
if "reset_code" not in st.session_state:
    st.session_state.reset_code = ""
if "reset_user" not in st.session_state:
    st.session_state.reset_user = ""
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "role" not in st.session_state:
    st.session_state.role = ""
if "branch_name" not in st.session_state:
    st.session_state.branch_name = ""
if "lang_key" not in st.session_state:
    st.session_state.lang_key = "vi"
if "dist_rows_count" not in st.session_state:
    st.session_state.dist_rows_count = 1
if "transfer_rows_count" not in st.session_state:
    st.session_state.transfer_rows_count = 1

LANG = {
    "vi": {
        "login_title": "Đăng Nhập Hệ Thống Quản Lý Kho & Chuỗi Nhà Hàng",
        "login_desc": "Vui lòng chọn ngôn ngữ, nhập **ID tài khoản** và **Mật khẩu** để tiếp tục.",
        "id_label": "ID tài khoản (Admin hoặc Chi nhánh):",
        "pwd_label": "Mật khẩu:",
        "btn_login": "Đăng Nhập",
        "title": "Hệ Thống Quản Lý Kho & Sơ Chế Chuỗi Nhà Hàng",
        "menu": "Chức Năng Hệ Thống",
        "m_overview": "Tổng Quan & Cảnh Báo Kho",
        "m_import": "Nhập Hàng Kho Tổng",
        "m_edit": "Sửa Tồn Kho Đầu Kỳ",
        "m_add": "Thêm Sản Phẩm Mới",
        "m_process": "Sơ Chế & Hao Hụt",
        "m_distribute": "Cấp Hàng Cho Chi Nhánh",
        "m_transfer": "Chuyển Hàng Giữa Chi Nhánh",
        "m_order": "Chi Nhánh Đặt Hàng Kho Tổng",
        "m_guide": "Hướng Dẫn Sử Dụng",
        "branch_restricted": "Bạn đang xem ở chế độ Chi nhánh. Một số tính năng quản lý thuộc về Kho Tổng.",
        "total_items": "Tổng số mặt hàng",
        "low_stock_warn": "Cảnh báo tồn kho thấp",
        "total_branches": "Tổng số chi nhánh hoạt động",
        "main_stock_table": "Bảng Tồn Kho Kho Tổng (Main Stock)",
        "import_desc": "Nhập số lượng hàng mới về từ nhà cung cấp vào kho tổng.",
        "sel_item_import": "Chọn sản phẩm cần nhập:",
        "import_qty": "Số lượng nhập thêm:",
        "import_source": "Nhà cung cấp / Nguồn gốc:",
        "btn_confirm_import": "Xác Nhận Nhập Hàng",
        "edit_desc": "Chỉnh sửa tồn kho đầu kỳ của sản phẩm.",
        "sel_item_edit": "Chọn sản phẩm cần sửa:",
        "new_opening": "Tồn kho đầu kỳ mới:",
        "pwd_confirm": "Nhập mật khẩu bảo mật để xác nhận:",
        "btn_update_stock": "Cập Nhật Tồn Kho",
        "add_desc": "Thêm mặt hàng mới vào danh mục kho tổng (Mã sản phẩm tự động sinh).",
        "item_id": "Mã sản phẩm (Tự động sinh):",
        "item_name": "Tên sản phẩm (ItemName):",
        "unit_main": "Đơn vị tính chính (Unit):",
        "opening_stock": "Tồn kho đầu kỳ:",
        "supplier": "Nguồn cung cấp / Nhà cung cấp:",
        "sub_unit_label": "Chọn đơn vị quy đổi (nếu có):",
        "sub_qty_label": "Nhập định mức cho 1 đơn vị chính (Số lượng/thể tích):",
        "btn_add_item": "Thêm Sản Phẩm",
        "process_desc": "Ghi nhận quá trình sơ chế nguyên liệu thô thành thành phẩm và hao hụt.",
        "proc_date": "Ngày sơ chế:",
        "batch_id": "Mã lô (BatchID):",
        "raw_material": "Nguyên liệu thô sử dụng:",
        "used_qty": "Số lượng nguyên liệu dùng:",
        "finished_product": "Tên thành phẩm thu được:",
        "produced_qty": "Số lượng thành phẩm thu được:",
        "waste": "Hao hụt / Phế phẩm:",
        "note": "Ghi chú chi tiết:",
        "btn_save_process": "Xác Nhận Lưu Sơ Chế",
        "dist_desc": "Cấp phát hàng hóa từ kho tổng cho các chi nhánh (có thể thêm nhiều sản phẩm bằng nút +).",
        "sel_branch": "Chọn chi nhánh nhận hàng:",
        "dist_item": "Chọn sản phẩm xuất kho:",
        "dist_qty": "Số lượng cấp phát:",
        "btn_add_row": "➕ Thêm sản phẩm khác",
        "btn_confirm_dist": "Xác Nhận Cấp Hàng",
        "transfer_desc": "Chuyển hàng nội bộ giữa các chi nhánh hoặc từ chi nhánh về kho (có thể chọn nhiều sản phẩm bằng nút +).",
        "from_unit": "Từ đơn vị gửi:",
        "to_unit": "Đến đơn vị nhận:",
        "trans_item": "Sản phẩm chuyển:",
        "trans_qty": "Số lượng chuyển:",
        "staff": "Nhân viên thực hiện:",
        "btn_confirm_trans": "Xác Nhận Chuyển Hàng",
        "order_desc_branch": "Giao diện đặt hàng dành cho chi nhánh:",
        "sel_product_order": "Chọn sản phẩm muốn đặt:",
        "order_qty": "Số lượng đặt:",
        "order_note": "Ghi chú thêm cho Kho Tổng:",
        "btn_send_order": "Gửi Yêu Cầu Đặt Hàng",
        "order_manage_title": "Quản lý tất cả đơn hàng từ các chi nhánh",
        "update_status_title": "Cập nhật trạng thái đơn hàng",
        "order_index_input": "Nhập số thứ tự dòng đơn hàng cần xử lý (bắt đầu từ 0):",
        "new_status_select": "Đổi trạng thái thành:",
        "btn_update_status": "Cập Nhật Trạng Thái Đơn",
        "guide_content": """### Hướng Dẫn Sử Dụng Hệ Thống
* **Quản trị viên (Admin):** Toàn quyền quản lý kho tổng, nhập hàng, thêm sản phẩm, duyệt đơn và phân bổ cấp hàng cho chi nhánh.
* **Chi nhánh (Branch):** Xem tổng quan tồn kho, chuyển hàng nội bộ và đặt hàng từ kho tổng."""
    },
    "en": {
        "login_title": "Restaurant Inventory & Chain Management Login",
        "login_desc": "Please select your language, enter your **Account ID** and **Password** to continue.",
        "id_label": "Account ID (Admin or Branch):",
        "pwd_label": "Password:",
        "btn_login": "Log In",
        "title": "Restaurant Inventory & Chain Management System",
        "menu": "System Menu",
        "m_overview": "Overview & Stock Alerts",
        "m_import": "Main Stock Import",
        "m_edit": "Edit Opening Stock",
        "m_add": "Add New Item",
        "m_process": "Processing Log",
        "m_distribute": "Branch Distribution",
        "m_transfer": "Inter-branch Transfer",
        "m_order": "Branch Order Request",
        "m_guide": "User Guide",
        "branch_restricted": "You are viewing in Branch mode. Management features belong to Main Inventory.",
        "total_items": "Total Items",
        "low_stock_warn": "Low Stock Alerts",
        "total_branches": "Active Branches",
        "main_stock_table": "Main Inventory Stock Table",
        "import_desc": "Import new items from suppliers into the main inventory.",
        "sel_item_import": "Select item to import:",
        "import_qty": "Import quantity:",
        "import_source": "Supplier / Origin:",
        "btn_confirm_import": "Confirm Import",
        "edit_desc": "Edit opening stock for products.",
        "sel_item_edit": "Select item to edit:",
        "new_opening": "New opening stock:",
        "pwd_confirm": "Enter security password to confirm:",
        "btn_update_stock": "Update Stock",
        "add_desc": "Add new item to main catalog (Auto-generated product ID).",
        "item_id": "Item ID (Auto):",
        "item_name": "Item Name:",
        "unit_main": "Main Unit:",
        "opening_stock": "Opening Stock:",
        "supplier": "Supplier / Source:",
        "sub_unit_label": "Select conversion unit (if any):",
        "sub_qty_label": "Enter quota for 1 main unit (Quantity/Volume):",
        "btn_add_item": "Add Product",
        "process_desc": "Log processing of raw materials into finished goods and waste.",
        "proc_date": "Processing Date:",
        "batch_id": "Batch ID:",
        "raw_material": "Raw Material Used:",
        "used_qty": "Quantity Used:",
        "finished_product": "Finished Product Name:",
        "produced_qty": "Quantity Produced:",
        "waste": "Waste / Loss:",
        "note": "Detailed Notes:",
        "btn_save_process": "Confirm & Save Processing",
        "dist_desc": "Distribute goods from main inventory to branches (add multiple items using + button).",
        "sel_branch": "Select receiving branch:",
        "dist_item": "Select item to distribute:",
        "dist_qty": "Distribution quantity:",
        "btn_add_row": "➕ Add another item",
        "btn_confirm_dist": "Confirm Distribution",
        "transfer_desc": "Transfer goods internally between branches or back to main inventory (add multiple items using + button).",
        "from_unit": "From Unit:",
        "to_unit": "To Unit:",
        "trans_item": "Transfer Item:",
        "trans_qty": "Transfer Quantity:",
        "staff": "Staff In Charge:",
        "btn_confirm_trans": "Confirm Transfer",
        "order_desc_branch": "Ordering interface for branch:",
        "sel_product_order": "Select product to order:",
        "order_qty": "Order Quantity:",
        "order_note": "Additional notes for Main Inventory:",
        "btn_send_order": "Submit Order Request",
        "order_manage_title": "Manage All Branch Orders",
        "update_status_title": "Update Order Status",
        "order_index_input": "Enter order row index to process (starts at 0):",
        "new_status_select": "Change status to:",
        "btn_update_status": "Update Order Status",
        "guide_content": """### User Guide
* **Administrator (Admin):** Full control over main stock, imports, order approvals, and branch distribution.
* **Branch:** View inventory, internal transfers, and place orders from main inventory."""
    },
    "hu": {
        "login_title": "Éttermi Készletkezelő Bejelentkezés",
        "login_desc": "Kérjük, adja meg a fiókazonosítót és a jelszót.",
        "id_label": "Fiók azonosító:",
        "pwd_label": "Jelszó:",
        "btn_login": "Bejelentkezés",
        "title": "Éttermi Készletkezelő és Lánc Rendszer",
        "menu": "Rendszer Menü",
        "m_overview": "Áttekintés és Készletriasztások",
        "m_import": "Központi Készlet Bevételezés",
        "m_edit": "Nyitókészlet Szerkesztése",
        "m_add": "Új Termék Hozzáadása",
        "m_process": "Feldolgozási Napló",
        "m_distribute": "Kiosztás Egységeknek",
        "m_transfer": "Egységek közötti átadás",
        "m_order": "Egységek Rendelése a Központból",
        "m_guide": "Használati Útmutató",
        "branch_restricted": "Fiók módban tekintettel. A kezelési funkciók a központi készlethez tartoznak.",
        "total_items": "Összes termék",
        "low_stock_warn": "Alacsony készlet riasztás",
        "total_branches": "Aktív egységek",
        "main_stock_table": "Központi Készlet Táblázat",
        "import_desc": "Új termékek bevételezése a központi raktárba.",
        "sel_item_import": "Válassza ki a terméket:",
        "import_qty": "Bevételezési mennyiség:",
        "import_source": "Beszállító / Forrás:",
        "btn_confirm_import": "Bevételezés Jóváhagyása",
        "edit_desc": "Nyitókészlet módosítása a termékekhez.",
        "sel_item_edit": "Szerkesztendő termék:",
        "new_opening": "Új nyitókészlet:",
        "pwd_confirm": "Adja meg a biztonsági jelszót a jóváhagyáshoz:",
        "btn_update_stock": "Készlet Frissítése",
        "add_desc": "Új termék hozzáadása a központi katalógushoz.",
        "item_id": "Termékazonosító (Auto):",
        "item_name": "Termék neve:",
        "unit_main": "Fő egység:",
        "opening_stock": "Nyitókészlet:",
        "supplier": "Beszállító:",
        "sub_unit_label": "Válasszon átváltási egységet (ha van):",
        "sub_qty_label": "Írja be az 1 fő egységre eső mennyiséget/térfogatot:",
        "btn_add_item": "Termék Hozzáadása",
        "process_desc": "Alapanyagok feldolgozásának rögzítése.",
        "proc_date": "Feldolgozás dátuma:",
        "batch_id": "Tételszám:",
        "raw_material": "Felhasznált alapanyag:",
        "used_qty": "Felhasznált mennyiség:",
        "finished_product": "Késztermék neve:",
        "produced_qty": "Előállított mennyiség:",
        "waste": "Hulladék / Veszteség:",
        "note": "Megjegyzések:",
        "btn_save_process": "Feldolgozás Mentése",
        "dist_desc": "Árukiosztás a központi raktárból az egységeknek.",
        "sel_branch": "Fogadó egység kiválasztása:",
        "dist_item": "Kiosztandó termék:",
        "dist_qty": "Kiosztási mennyiség:",
        "btn_add_row": "➕ További termék hozzáadása",
        "btn_confirm_dist": "Kiosztás Jóváhagyása",
        "transfer_desc": "Belső árumozgatás.",
        "from_unit": "Küldő egység:",
        "to_unit": "Fogadó egység:",
        "trans_item": "Átidegenítendő termék:",
        "trans_qty": "Mennyiség:",
        "staff": "Felelős munkatárs:",
        "btn_confirm_trans": "Átutalás Jóváhagyása",
        "order_desc_branch": "Rendelési felület az egység számára:",
        "sel_product_order": "Rendelni kívánt termék:",
        "order_qty": "Rendelési mennyiség:",
        "order_note": "Megjegyzések a Központnak:",
        "btn_send_order": "Rendelés Beküldése",
        "order_manage_title": "Összes egységi rendelés kezelése",
        "update_status_title": "Rendelési státusz frissítése",
        "order_index_input": "Írja be a rendelés sorszámát (0-tól):",
        "new_status_select": "Státusz módosítása:",
        "btn_update_status": "Státusz Frissítése",
        "guide_content": """### Használati Útmutató
* **Rendszergazda (Admin):** Teljes körű vezérlés a központi raktár, bevételezések és kiosztások felett.
* **Egység (Branch):** Összesített készletek megtekintése, belső átadások és rendelések leadása."""
    }
}

# --- GIAO DIỆN ĐĂNG NHẬP ---
if not st.session_state.logged_in:
    login_lang_choice = st.selectbox("Chọn ngôn ngữ / Language / Nyelv:", ["Tiếng Việt", "English", "Magyar"], index=0)
    login_lang_key = "vi" if login_lang_choice == "Tiếng Việt" else ("en" if login_lang_choice == "English" else "hu")
    TL = LANG[login_lang_key]

    st.title(TL["login_title"])
    st.markdown(TL["login_desc"])
    
    tab_login, tab_forgot = st.tabs(["Đăng Nhập" if login_lang_key=="vi" else ("Log In" if login_lang_key=="en" else "Bejelentkezés"), "Quên Mật Khẩu" if login_lang_key=="vi" else ("Forgot Password" if login_lang_key=="en" else "Elfelejtett Jelszó")])
    
    with tab_login:
        with st.form("login_form"):
            input_id = st.text_input(TL["id_label"])
            input_pwd = st.text_input(TL["pwd_label"], type="password")
            submitted_login = st.form_submit_button(TL["btn_login"])
            
            if submitted_login:
                clean_id = input_id.strip().lower()
                if clean_id in st.session_state.passwords and input_pwd == st.session_state.passwords[clean_id]:
                    st.session_state.logged_in = True
                    if clean_id in ["nuonuo", "heni", "admin"]:
                        st.session_state.role = "Admin"
                        st.session_state.branch_name = ""
                    else:
                        st.session_state.role = "Branch"
                        branch_names_map = {
                            "shibuya": "Shibuya",
                            "geisha.baross": "Little Geisha Baross",
                            "geisha.corvin": "Little Geisha Corvin",
                            "urbn": "URBN.Station",
                            "matchy": "Matchy"
                        }
                        st.session_state.branch_name = branch_names_map.get(clean_id, clean_id)
                    st.session_state.lang_key = login_lang_key
                    st.success("Đăng nhập thành công / Login successful!")
                    st.rerun()
                else:
                    st.error("Sai ID hoặc mật khẩu! / Incorrect ID or Password!")

    with tab_forgot:
        st.markdown("Khôi phục mật khẩu qua email: **nuo26420gmail.com**")
        if st.session_state.reset_step == 1:
            with st.form("forgot_step1"):
                f_user = st.text_input("Nhập ID tài khoản / Enter Account ID:")
                btn_send = st.form_submit_button("Gửi mã xác nhận / Send Code")
                if btn_send:
                    clean_f_user = f_user.strip().lower()
                    if clean_f_user in st.session_state.passwords:
                        generated_code = str(random.randint(100000, 999999))
                        st.session_state.reset_code = generated_code
                        st.session_state.reset_user = clean_f_user
                        st.session_state.reset_step = 2
                        st.success(f"Mã xác nhận đã gửi về email! (Mã test: {generated_code})")
                        st.rerun()
                    else:
                        st.error("ID tài khoản không tồn tại!")
        elif st.session_state.reset_step == 2:
            with st.form("forgot_step2"):
                st.info(f"Đang khôi phục cho: **{st.session_state.reset_user}**")
                input_code = st.text_input("Nhập mã xác nhận (6 chữ số):")
                new_pwd1 = st.text_input("Mật khẩu mới:", type="password")
                new_pwd2 = st.text_input("Nhập lại mật khẩu mới:", type="password")
                btn_confirm = st.form_submit_button("Xác Nhận Đổi Mật Khẩu")
                if btn_confirm:
                    if input_code.strip() != st.session_state.reset_code:
                        st.error("Mã xác nhận không chính xác!")
                    elif not new_pwd1 or new_pwd1 != new_pwd2:
                        st.error("Mật khẩu mới không khớp!")
                    else:
                        st.session_state.passwords[st.session_state.reset_user] = new_pwd1
                        st.session_state.reset_step = 1
                        st.success("Đổi mật khẩu thành công! Vui lòng đăng nhập lại.")
        if st.button("Hủy / Nhập lại"):
            st.session_state.reset_step = 1
            st.rerun()
    st.stop()

# --- HÀM XỬ LÝ DỮ LIỆU ---
def load_data():
    if os.path.exists(DATA_FILE):
        try:
            main_stock = pd.read_excel(DATA_FILE, sheet_name="MainStock")
            processing = pd.read_excel(DATA_FILE, sheet_name="ProcessingLog")
        except:
            main_stock = pd.DataFrame(columns=["ItemID", "ItemName", "Unit", "OpeningStock", "Import", "BranchExport", "ProcessingExport", "Source", "SubUnit", "SubQuantity", "TotalConverted"])
            processing = pd.DataFrame(columns=["Date", "BatchID", "RawMaterial", "UsedQuantity", "FinishedProduct", "ProducedQuantity", "WasteLoss", "Note"])
        
        for col in ["OpeningStock", "Import", "BranchExport", "ProcessingExport"]:
            if col in main_stock.columns:
                main_stock[col] = pd.to_numeric(main_stock[col], errors="coerce").fillna(0.0)
        if "Source" not in main_stock.columns:
            main_stock["Source"] = "Nhà cung cấp chính"
        if "SubUnit" not in main_stock.columns:
            main_stock["SubUnit"] = ""
        if "SubQuantity" not in main_stock.columns:
            main_stock["SubQuantity"] = 0.0
        if "TotalConverted" not in main_stock.columns:
            main_stock["TotalConverted"] = ""
        return main_stock, processing
    else:
        main_stock = pd.DataFrame(columns=["ItemID", "ItemName", "Unit", "OpeningStock", "Import", "BranchExport", "ProcessingExport", "Source", "SubUnit", "SubQuantity", "TotalConverted"])
        processing = pd.DataFrame(columns=["Date", "BatchID", "RawMaterial", "UsedQuantity", "FinishedProduct", "ProducedQuantity", "WasteLoss", "Note"])
        save_data(main_stock, processing)
        return main_stock, processing

def save_data(main_stock, processing):
    if "Source" not in main_stock.columns:
        main_stock["Source"] = "Nhà cung cấp chính"
    if "SubUnit" not in main_stock.columns:
        main_stock["SubUnit"] = ""
    if "SubQuantity" not in main_stock.columns:
        main_stock["SubQuantity"] = 0.0
    if "TotalConverted" not in main_stock.columns:
        main_stock["TotalConverted"] = ""
    with pd.ExcelWriter(DATA_FILE, engine="openpyxl") as writer:
        main_stock.to_excel(writer, sheet_name="MainStock", index=False)
        processing.to_excel(writer, sheet_name="ProcessingLog", index=False)
    format_excel_file(DATA_FILE)

if not os.path.exists(EXPORT_FILE):
    pd.DataFrame(columns=["ExportDate", "Branch", "ItemName", "Unit", "Quantity", "Sender", "Receiver"]).to_csv(EXPORT_FILE, index=False)

if not os.path.exists(TRANSFER_FILE):
    pd.DataFrame(columns=["TransferDate", "FromBranch", "ToBranch", "ItemName", "Unit", "Quantity", "Staff"]).to_csv(TRANSFER_FILE, index=False)

if not os.path.exists(ORDER_FILE):
    pd.DataFrame(columns=["OrderDate", "Branch", "ItemName", "Unit", "Quantity", "Status", "Note"]).to_csv(ORDER_FILE, index=False)

def calculate_closing_stock(df):
    if df.empty:
        df["ClosingStock"] = []
        return df
    for col in ["OpeningStock", "Import", "BranchExport", "ProcessingExport"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    if "Source" not in df.columns:
        df["Source"] = "Nhà cung cấp chính"
    if "SubUnit" not in df.columns:
        df["SubUnit"] = ""
    if "SubQuantity" not in df.columns:
        df["SubQuantity"] = 0.0
    if "TotalConverted" not in df.columns:
        df["TotalConverted"] = ""
    df["ClosingStock"] = df["OpeningStock"] + df["Import"] - df["BranchExport"] - df["ProcessingExport"]
    return df

def format_excel_file(file_path):
    wb = None
    try:
        wb = openpyxl.load_workbook(file_path)
        thin_border = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'), top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            if ws.max_row < 1: continue
            header_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            data_font = Font(name="Calibri", size=12)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 20
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value: max_length = max(max_length, len(str(cell.value)))
                    except: pass
                ws.column_dimensions[col_letter].width = max(max_length + 4, 12)
        wb.save(file_path)
    except Exception:
        pass
    finally:
        if wb is not None:
            try:
                wb.close()
            except:
                pass

# --- THANH SIDEBAR ---
current_lang_key = st.session_state.get("lang_key", "vi")
T = LANG[current_lang_key]

st.sidebar.header("Ngôn Ngữ / Language")
selected_lang_ui = st.sidebar.selectbox(
    "Chọn ngôn ngữ hiển thị:", 
    ["Tiếng Việt", "English", "Magyar"], 
    index=0 if current_lang_key == "vi" else (1 if current_lang_key == "en" else 2),
    key="ui_lang_select"
)
new_lang_key = "vi" if selected_lang_ui == "Tiếng Việt" else ("en" if selected_lang_ui == "English" else "hu")
if new_lang_key != st.session_state.lang_key:
    st.session_state.lang_key = new_lang_key
    st.rerun()

T = LANG[st.session_state.lang_key]

st.sidebar.markdown("---")
st.sidebar.markdown(f"Tài khoản: `{st.session_state.role}`")
if st.session_state.role == "Branch":
    st.sidebar.markdown(f"Chi nhánh: `{st.session_state.branch_name}`")

if st.sidebar.button("Đăng Xuất / Log Out", use_container_width=True):
    st.session_state.logged_in = False
    st.session_state.role = ""
    st.session_state.branch_name = ""
    st.rerun()

st.sidebar.markdown("---")

pending_order_count = 0
if os.path.exists(ORDER_FILE):
    try:
        df_check_order = pd.read_csv(ORDER_FILE)
        if not df_check_order.empty and "Status" in df_check_order.columns:
            pending_order_count = len(df_check_order[df_check_order["Status"].str.contains("chờ", case=False, na=False)])
    except:
        pass

st.sidebar.title(T["menu"])

order_menu_label = T["m_order"]
if st.session_state.role == "Admin" and pending_order_count > 0:
    order_menu_label = f"🔴 {T['m_order']} ({pending_order_count} chờ duyệt)"

if st.session_state.role == "Admin":
    menu_options = {
        T["m_overview"]: "overview",
        T["m_import"]: "import",
        T["m_edit"]: "edit",
        T["m_add"]: "add",
        T["m_process"]: "process",
        T["m_distribute"]: "distribute",
        T["m_transfer"]: "transfer",
        order_menu_label: "order",
        T["m_guide"]: "guide"
    }
else:
    menu_options = {
        T["m_overview"]: "overview",
        T["m_transfer"]: "transfer",
        T["m_order"]: "order",
        T["m_guide"]: "guide"
    }

selected_menu_label = st.sidebar.radio("Chọn chức năng:", list(menu_options.keys()))
choice = menu_options[selected_menu_label]

st.title(T["title"])

main_stock_df, processing_df = load_data()
main_stock_df = calculate_closing_stock(main_stock_df)

admin_only_choices = ["import", "edit", "add", "process", "distribute"]
if st.session_state.role == "Branch" and choice in admin_only_choices:
    st.warning(T["branch_restricted"])
    st.stop()

# --- XỬ LÝ CÁC TÍNH NĂNG CHÍNH ---
if choice == "overview":
    st.subheader(T["m_overview"])
    total_items = len(main_stock_df)
    low_stock_items = len(main_stock_df[main_stock_df["ClosingStock"] <= 5])
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(label=T["total_items"], value=total_items)
    with col2:
        st.metric(label=T["low_stock_warn"], value=low_stock_items, delta="Ổn định" if low_stock_items == 0 else "Chú ý", delta_color="inverse")
    with col3:
        st.metric(label=T["total_branches"], value=5)
        
    st.markdown("---")
    st.markdown(f"### {T['main_stock_table']}")
    st.dataframe(main_stock_df, use_container_width=True)

elif choice == "import":
    st.subheader(T["m_import"])
    st.markdown(T["import_desc"])
    with st.form("import_form"):
        item_list = main_stock_df["ItemName"].tolist() if not main_stock_df.empty else []
        selected_item = st.selectbox(T["sel_item_import"], item_list)
        import_qty = st.number_input(T["import_qty"], min_value=0.0, step=1.0)
        import_source = st.text_input(T["import_source"], value="Nhà cung cấp chính")
        submitted_import = st.form_submit_button(T["btn_confirm_import"])
        if submitted_import:
            if selected_item:
                idx = main_stock_df[main_stock_df["ItemName"] == selected_item].index
                if not idx.empty:
                    main_stock_df.loc[idx, "Import"] += import_qty
                    main_stock_df.loc[idx, "Source"] = import_source
                    save_data(main_stock_df, processing_df)
                    st.success("Thành công!")
                    st.rerun()

elif choice == "edit":
    st.subheader(T["m_edit"])
    st.markdown(T["edit_desc"])
    with st.form("edit_stock_form"):
        item_list = main_stock_df["ItemName"].tolist() if not main_stock_df.empty else []
        selected_item = st.selectbox(T["sel_item_edit"], item_list)
        new_opening = st.number_input(T["new_opening"], min_value=0.0, step=1.0)
        pwd_input = st.text_input(T["pwd_confirm"], type="password")
        submitted_edit = st.form_submit_button(T["btn_update_stock"])
        if submitted_edit:
            if pwd_input != SECRET_ACTION_PWD:
                st.error("Sai mật khẩu xác nhận!")
            elif selected_item:
                idx = main_stock_df[main_stock_df["ItemName"] == selected_item].index
                if not idx.empty:
                    main_stock_df.loc[idx, "OpeningStock"] = new_opening
                    save_data(main_stock_df, processing_df)
                    st.success("Cập nhật thành công!")
                    st.rerun()

elif choice == "add":
    st.subheader(T["m_add"])
    st.markdown(T["add_desc"])
    
    if main_stock_df.empty:
        auto_id = "SP001"
    else:
        try:
            numeric_parts = main_stock_df["ItemID"].str.extract(r'(\d+)')[0].dropna().astype(int)
            next_num = numeric_parts.max() + 1 if not numeric_parts.empty else len(main_stock_df) + 1
            auto_id = f"SP{next_num:03d}"
        except:
            auto_id = f"SP{len(main_stock_df)+1:03d}"

    with st.form("add_item_form"):
        new_id = st.text_input(T["item_id"], value=auto_id, disabled=True)
        new_name = st.text_input(T["item_name"])
        unit_options = ["Kg", "g", "L (Lít)", "ml", "Can", "Bottle (Chai)", "Thùng", "Gói", "Bịch", "Lon", "Hộp", "Cái"]
        new_unit = st.selectbox(T["unit_main"], unit_options)
        new_opening = st.number_input(T["opening_stock"], min_value=0.0, step=1.0)
        new_source = st.text_input(T["supplier"], value="Nhà cung cấp chính")
        
        st.markdown("---")
        sub_unit_options = ["Không", "ml", "g", "L", "Kg", "Lon", "Chai", "Gói", "Bịch", "Thùng", "Hộp", "Cái"]
        sub_unit_choice = st.selectbox(T["sub_unit_label"], sub_unit_options)
        sub_qty_per_unit = st.number_input(T["sub_qty_label"], min_value=0.0, step=1.0)
        
        submitted_add = st.form_submit_button(T["btn_add_item"])
        
        if submitted_add:
            if new_name.strip() == "":
                st.error("Tên sản phẩm không được để trống!")
            elif not main_stock_df[main_stock_df["ItemID"] == auto_id].empty:
                st.error(f"Mã sản phẩm đã tồn tại!")
            else:
                calculated_total_str = ""
                if sub_unit_choice != "Không" and sub_qty_per_unit > 0:
                    total_val = new_opening * sub_qty_per_unit
                    calculated_total_str = f"{total_val:,.0f} {sub_unit_choice}"
                
                new_row = {
                    "ItemID": auto_id,
                    "ItemName": new_name,
                    "Unit": new_unit,
                    "OpeningStock": new_opening,
                    "Import": 0.0,
                    "BranchExport": 0.0,
                    "ProcessingExport": 0.0,
                    "Source": new_source,
                    "SubUnit": sub_unit_choice if sub_unit_choice != "Không" else "",
                    "SubQuantity": sub_qty_per_unit if sub_qty_per_unit > 0 else 0.0,
                    "TotalConverted": calculated_total_str
                }
                main_stock_df = pd.concat([main_stock_df, pd.DataFrame([new_row])], ignore_index=True)
                save_data(main_stock_df, processing_df)
                st.success(f"Thêm thành công sản phẩm **{new_name}**!")
                st.rerun()

elif choice == "process":
    st.subheader(T["m_process"])
    st.markdown(T["process_desc"])
    with st.form("process_form"):
        p_date = st.date_input(T["proc_date"], datetime.now())
        p_batch = st.text_input(T["batch_id"], value=f"BATCH-{datetime.now().strftime('%Y%m%d')}")
        item_list = main_stock_df["ItemName"].tolist() if not main_stock_df.empty else []
        p_raw = st.selectbox(T["raw_material"], item_list)
        p_used_qty = st.number_input(T["used_qty"], min_value=0.0, step=1.0)
        p_finished = st.text_input(T["finished_product"])
        p_produced_qty = st.number_input(T["produced_qty"], min_value=0.0, step=1.0)
        p_waste = st.number_input(T["waste"], min_value=0.0, step=1.0)
        p_note = st.text_area(T["note"])
        submitted_process = st.form_submit_button(T["btn_save_process"])
        if submitted_process:
            idx = main_stock_df[main_stock_df["ItemName"] == p_raw].index
            if not idx.empty:
                main_stock_df.loc[idx, "ProcessingExport"] += p_used_qty
                new_log = {
                    "Date": p_date.strftime("%Y-%m-%d"),
                    "BatchID": p_batch,
                    "RawMaterial": p_raw,
                    "UsedQuantity": p_used_qty,
                    "FinishedProduct": p_finished,
                    "ProducedQuantity": p_produced_qty,
                    "WasteLoss": p_waste,
                    "Note": p_note
                }
                processing_df = pd.concat([processing_df, pd.DataFrame([new_log])], ignore_index=True)
                save_data(main_stock_df, processing_df)
                st.success("Đã ghi nhận sơ chế thành công!")
                st.rerun()

elif choice == "distribute":
    st.subheader(T["m_distribute"])
    st.markdown(T["dist_desc"])
    
    with st.form("distribute_form"):
        d_branch = st.selectbox(T["sel_branch"], ["Shibuya", "Little Geisha Baross", "Little Geisha Corvin", "URBN.Station", "Matchy"])
        
        item_list_dist = main_stock_df["ItemName"].tolist() if not main_stock_df.empty else []
        selected_items_data = []
        
        for i in range(st.session_state.dist_rows_count):
            st.markdown(f"**Sản phẩm #{i+1}**")
            col_i1, col_i2 = st.columns([2, 1])
            with col_i1:
                d_item = st.selectbox(T["dist_item"], item_list_dist, key=f"dist_item_{i}") if item_list_dist else None
            with col_i2:
                d_qty = st.number_input(T["dist_qty"], min_value=0.0, step=1.0, key=f"dist_qty_{i}")
            selected_items_data.append((d_item, d_qty))
            
        submitted_add_row = st.form_submit_button(T["btn_add_row"])
        submitted_dist = st.form_submit_button(T["btn_confirm_dist"])
        
        if submitted_add_row:
            st.session_state.dist_rows_count += 1
            st.rerun()
            
        if submitted_dist:
            valid_any = False
            exp_df = pd.read_csv(EXPORT_FILE)
            for d_item, d_qty in selected_items_data:
                if d_item and d_qty > 0:
                    valid_any = True
                    idx = main_stock_df[main_stock_df["ItemName"] == d_item].index
                    if not idx.empty:
                        main_stock_df.loc[idx, "BranchExport"] += d_qty
                        
                    export_record = {
                        "ExportDate": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "Branch": d_branch,
                        "ItemName": d_item,
                        "Unit": "Kg/Thùng",
                        "Quantity": d_qty,
                        "Sender": "Kho Tổng",
                        "Receiver": d_branch
                    }
                    exp_df = pd.concat([exp_df, pd.DataFrame([export_record])], ignore_index=True)
            
            if valid_any:
                exp_df.to_csv(EXPORT_FILE, index=False)
                save_data(main_stock_df, processing_df)
                st.session_state.dist_rows_count = 1
                st.success(f"Đã cấp thành công hàng cho chi nhánh **{d_branch}**!")
                st.rerun()
            else:
                st.warning("Vui lòng chọn ít nhất một sản phẩm với số lượng lớn hơn 0.")

    st.markdown("---")
    st.markdown("### 📋 Lịch Sử Cấp Hàng Cho Chi Nhánh")
    if os.path.exists(EXPORT_FILE):
        df_exp_hist = pd.read_csv(EXPORT_FILE)
        if df_exp_hist.empty:
            st.info("Chưa có lịch sử cấp hàng nào.")
        else:
            st.dataframe(df_exp_hist, use_container_width=True)
    else:
        st.info("Chưa có lịch sử cấp hàng nào.")

elif choice == "transfer":
    st.subheader(T["m_transfer"])
    st.markdown(T["transfer_desc"])
    
    branch_list = ["Shibuya", "Little Geisha Baross", "Little Geisha Corvin", "URBN.Station", "Matchy", "Kho Tổng"]
    
    with st.form("transfer_form"):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            t_from = st.selectbox(T["from_unit"], branch_list)
        with col_f2:
            t_to = st.selectbox(T["to_unit"], branch_list, index=1 if len(branch_list) > 1 else 0)
            
        t_staff = st.text_input(T["staff"])
        
        item_list_trans = main_stock_df["ItemName"].tolist() if not main_stock_df.empty else []
        selected_trans_data = []
        
        for i in range(st.session_state.transfer_rows_count):
            st.markdown(f"**Sản phẩm chuyển #{i+1}**")
            col_t1, col_t2 = st.columns([2, 1])
            with col_t1:
                t_item = st.selectbox(T["trans_item"], item_list_trans, key=f"trans_item_{i}") if item_list_trans else None
            with col_t2:
                t_qty = st.number_input(T["trans_qty"], min_value=0.0, step=1.0, key=f"trans_qty_{i}")
            selected_trans_data.append((t_item, t_qty))
            
        submitted_add_trans_row = st.form_submit_button(T["btn_add_row"])
        submitted_trans = st.form_submit_button(T["btn_confirm_trans"])
        
        if submitted_add_trans_row:
            st.session_state.transfer_rows_count += 1
            st.rerun()
            
        if submitted_trans:
            if t_from == t_to:
                st.error("Đơn vị gửi và nhận không được trùng nhau!")
            else:
                valid_any_trans = False
                trans_df = pd.read_csv(TRANSFER_FILE)
                for t_item, t_qty in selected_trans_data:
                    if t_item and t_qty > 0:
                        valid_any_trans = True
                        trans_record = {
                            "TransferDate": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "FromBranch": t_from,
                            "ToBranch": t_to,
                            "ItemName": t_item,
                            "Unit": "Kg/Thùng",
                            "Quantity": t_qty,
                            "Staff": t_staff
                        }
                        trans_df = pd.concat([trans_df, pd.DataFrame([trans_record])], ignore_index=True)
                
                if valid_any_trans:
                    trans_df.to_csv(TRANSFER_FILE, index=False)
                    st.session_state.transfer_rows_count = 1
                    st.success("Chuyển hàng nội bộ thành công!")
                    st.rerun()
                else:
                    st.warning("Vui lòng chọn ít nhất một sản phẩm với số lượng lớn hơn 0.")

    st.markdown("---")
    st.markdown("### 📋 Lịch Sử Chuyển Hàng Giữa Các Chi Nhánh")
    if os.path.exists(TRANSFER_FILE):
        df_trans_hist = pd.read_csv(TRANSFER_FILE)
        if df_trans_hist.empty:
            st.info("Chưa có lịch sử chuyển hàng nội bộ nào.")
        else:
            st.dataframe(df_trans_hist, use_container_width=True)
    else:
        st.info("Chưa có lịch sử chuyển hàng nội bộ nào.")

elif choice.startswith("🔴") or choice.startswith(T["m_order"]) or choice == "order":
    st.subheader(T["m_order"])
    if st.session_state.role == "Branch":
        st.markdown(f"{T['order_desc_branch']} **{st.session_state.branch_name}**")
        with st.form("branch_order_form"):
            order_item = st.selectbox(T["sel_product_order"], main_stock_df["ItemName"].tolist() if not main_stock_df.empty else [])
            order_qty = st.number_input(T["order_qty"], min_value=1.0, step=1.0)
            order_note = st.text_area(T["order_note"])
            submitted_order = st.form_submit_button(T["btn_send_order"])
            if submitted_order:
                new_order = {
                    "OrderDate": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Branch": st.session_state.branch_name,
                    "ItemName": order_item,
                    "Unit": "Kg/Thùng",
                    "Quantity": order_qty,
                    "Status": "Đang chờ duyệt",
                    "Note": order_note
                }
                order_df = pd.read_csv(ORDER_FILE)
                order_df = pd.concat([order_df, pd.DataFrame([new_order])], ignore_index=True)
                order_df.to_csv(ORDER_FILE, index=False)
                st.success("Gửi yêu cầu đặt hàng thành công!")
                st.rerun()

        st.markdown("---")
        st.markdown("### 📋 Lịch Sử Đặt Hàng Của Chi Nhánh")
        if os.path.exists(ORDER_FILE):
            df_ord_hist = pd.read_csv(ORDER_FILE)
            df_branch_ord = df_ord_hist[df_ord_hist["Branch"] == st.session_state.branch_name]
            if df_branch_ord.empty:
                st.info("Bạn chưa có đơn đặt hàng nào.")
            else:
                st.dataframe(df_branch_ord, use_container_width=True)
        else:
            st.info("Chưa có đơn đặt hàng nào.")
    else:
        if pending_order_count > 0:
            st.error(f"🚨 CẢNH BÁO: Hiện đang có **{pending_order_count}** đơn hàng mới từ các chi nhánh đang chờ duyệt!")
            
        st.markdown(f"### {T['order_manage_title']}")
        order_df = pd.read_csv(ORDER_FILE)
        if order_df.empty:
            st.info("Chưa có đơn hàng nào.")
        else:
            st.dataframe(order_df, use_container_width=True)
            st.markdown("---")
            st.markdown(f"#### {T['update_status_title']}")
            with st.form("update_order_form"):
                order_idx = st.number_input(T["order_index_input"], min_value=0, max_value=max(0, len(order_df)-1), step=1)
                new_status = st.selectbox(T["new_status_select"], ["Đang chờ duyệt", "Đã duyệt", "Từ chối"])
                submitted_update = st.form_submit_button(T["btn_update_status"])
                if submitted_update:
                    order_df.loc[order_idx, "Status"] = new_status
                    order_df.to_csv(ORDER_FILE, index=False)
                    st.success("Cập nhật trạng thái thành công!")
                    st.rerun()

elif choice == "guide":
    st.subheader(T["m_guide"])
    st.markdown(T["guide_content"])